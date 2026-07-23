"""Build a tiny workbook, load+normalize it, write output, reload and assert."""

from pathlib import Path

import openpyxl

from po_classifier import decide as decide_mod
from po_classifier import io_excel, pivots, score_rules
from po_classifier.kpmg_ref import load_taxonomy

CFG = {
    "input": {
        "sheets": ["2025"],
        "dedupe": "prefer_consolidated",
        "header_detect": "auto",
        "columns": {
            "po": ["PO Number"],
            "supplier": ["Supplier"],
            "category": ["Category"],
            "currency": ["Currency"],
            "amount": ["Amount"],
        },
        "fx_sheet": "References",
    },
    "output": {
        "sheet_scored": "Scored Full List",
        "sheet_kept": "Filtered Kept",
        "sheet_review": "Review Queue",
        "sheet_pivots": "Pivot Summaries",
    },
    "scoring": {"keep_threshold": 0.55, "review_threshold": 0.6},
}


def _make_input(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2025"
    ws.append(["Publication of Purchase Orders 2025", None, None, None, None])  # banner
    ws.append(["PO Number", "Supplier", "Category", "Currency", "Amount"])
    ws.append([1, "ACCENTURE LTD", "COMPUTER/SERVICES", "EUR", 54450.0])
    ws.append([2, "FUEL CO", "FUELS", "USD", 1000.0])
    ref = wb.create_sheet("References")
    ref.append(["FX Rates", None])
    ref.append(["Currency", "EUR per 1 unit"])
    ref.append(["USD", 0.8862])
    wb.save(path)


def test_round_trip(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    _make_input(inp)

    rows = io_excel.load_rows(inp, CFG)
    assert len(rows) == 2
    # Banner row must have been skipped; USD converted.
    usd_row = next(r for r in rows if r.currency == "USD")
    assert abs(usd_row.amount_eur - 886.2) < 0.01

    tax = load_taxonomy(Path(__file__).parent.parent / "po_classifier" / "taxonomy.yaml")
    rule_scores = [score_rules.score_row(r, tax) for r in rows]
    decisions = decide_mod.decide(rows, rule_scores, {}, 0.55, 0.6)
    tables = pivots.build_pivots(decisions)
    io_excel.write_output(out, decisions, tables, CFG)

    wb = openpyxl.load_workbook(out)
    assert set(["Scored Full List", "Filtered Kept", "Review Queue", "Pivot Summaries"]).issubset(
        set(wb.sheetnames)
    )
    scored = wb["Scored Full List"]
    assert scored.max_row == 3  # header + 2 data rows
    # Accenture computer services should be kept; fuel discarded.
    kept = wb["Filtered Kept"]
    kept_suppliers = [kept.cell(row=r, column=2).value for r in range(2, kept.max_row + 1)]
    assert "ACCENTURE LTD" in kept_suppliers
    assert "FUEL CO" not in kept_suppliers
