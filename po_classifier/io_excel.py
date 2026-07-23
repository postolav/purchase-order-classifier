"""Excel I/O: load & normalize the raw workbook; write the 4-sheet output.

The input files drift in layout between quarters, so column mapping is driven by
config aliases and the header row is auto-detected (banner/title rows are skipped).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .fx import build_fx_table, parse_amount, to_eur
from .models import Decision, PORow


def _norm_label(s) -> str:
    return " ".join(str(s or "").strip().lower().split())


def _looks_like_year(name: str) -> bool:
    return bool(re.fullmatch(r"\d{4}", name.strip()))


def _find_header_row(ws, po_aliases: List[str], max_scan: int = 8):
    """Return (header_row_index_1based, {canonical_or_label: col_idx}) by scanning
    the first rows for one that contains a PO-number-like header."""
    po_norm = {_norm_label(a) for a in po_aliases}
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for i, row in enumerate(rows, start=1):
        labels = {_norm_label(v) for v in row if v is not None}
        if labels & po_norm:
            header = {}
            for col_idx, v in enumerate(row):
                if v is not None and _norm_label(v):
                    header[_norm_label(v)] = col_idx
            return i, header
    return None, {}


def _map_columns(header: Dict[str, int], columns_cfg: Dict[str, List[str]]) -> Dict[str, int]:
    """Map canonical field -> column index using the first matching alias."""
    mapping: Dict[str, int] = {}
    for canonical, aliases in columns_cfg.items():
        for alias in aliases:
            key = _norm_label(alias)
            if key in header:
                mapping[canonical] = header[key]
                break
    return mapping


def _pick_amount(values, col_map, currency_val) -> Optional[float]:
    """Return the numeric amount, tolerating the known layout quirk where the
    amount lands in a 'Currency'-labelled column (currency code sits elsewhere)."""
    idx = col_map.get("amount")
    amt = parse_amount(values[idx]) if idx is not None and idx < len(values) else None
    if amt is not None:
        return amt
    # Fallback: if the mapped 'currency' column actually holds a number, use it.
    cidx = col_map.get("currency")
    if cidx is not None and cidx < len(values):
        alt = parse_amount(values[cidx])
        if alt is not None and not _is_currency_code(currency_val):
            return alt
    return None


def _is_currency_code(v) -> bool:
    return isinstance(v, str) and bool(re.fullmatch(r"[A-Za-z]{3}", v.strip()))


def load_rows(path: str | Path, cfg: dict) -> List[PORow]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    input_cfg = cfg["input"]
    columns_cfg = input_cfg["columns"]

    # FX table.
    fx = {"EUR": 1.0}
    fx_sheet = input_cfg.get("fx_sheet")
    if fx_sheet and fx_sheet in wb.sheetnames:
        fx = build_fx_table(wb[fx_sheet].iter_rows(values_only=True))

    # Decide which sheets to read.
    requested = list(input_cfg.get("sheets", []))
    present = [s for s in requested if s in wb.sheetnames]
    if input_cfg.get("dedupe") == "prefer_consolidated":
        years = [s for s in present if _looks_like_year(s)]
        sheets = [years[0]] if years else present
    else:
        sheets = present
    if not sheets:
        raise ValueError(f"None of the requested sheets {requested} found in {path}")

    rows: List[PORow] = []
    row_counter = 0
    for sheet in sheets:
        ws = wb[sheet]
        hdr_row, header = _find_header_row(ws, columns_cfg["po"])
        if hdr_row is None:
            continue
        col_map = _map_columns(header, columns_cfg)
        for values in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if values is None or all(v is None for v in values):
                continue

            def get(field):
                idx = col_map.get(field)
                return values[idx] if idx is not None and idx < len(values) else None

            po = get("po")
            supplier = get("supplier")
            if po is None and supplier is None:
                continue
            currency_val = get("currency")
            currency = currency_val if _is_currency_code(currency_val) else "EUR"
            amount = _pick_amount(values, col_map, currency_val)
            amount_eur, flagged = to_eur(amount, currency, fx)

            row_counter += 1
            rows.append(
                PORow(
                    po=str(po).strip() if po is not None else "",
                    supplier=str(supplier).strip() if supplier is not None else "",
                    category=str(get("category") or "").strip(),
                    currency=currency,
                    amount=amount,
                    amount_eur=amount_eur,
                    source_sheet=sheet,
                    row_index=row_counter,
                    fx_flag=flagged,
                )
            )
    wb.close()
    return rows


# --------------------------------------------------------------------------- write


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 60)


def write_output(path: str | Path, decisions: List[Decision], pivots: dict, cfg: dict) -> None:
    out_cfg = cfg["output"]
    wb = openpyxl.Workbook()

    # Sheet 1 - Scored Full List
    ws = wb.active
    ws.title = out_cfg["sheet_scored"]
    scored_hdr = [
        "PO Number", "Supplier", "Raw Category", "Currency", "Amount", "Amount (EUR)",
        "KPMG Category", "Similarity", "Confidence", "Decision", "Needs Review",
        "Reason", "Source Stage", "FX Flag",
    ]
    ws.append(scored_hdr)
    for d in decisions:
        r, s = d.row, d.score
        ws.append([
            r.po, r.supplier, r.category, r.currency, r.amount, r.amount_eur,
            s.kpmg_category or "", round(s.similarity, 3), round(s.confidence, 3),
            d.decision, "Y" if d.needs_review else "", s.reason, s.source_stage,
            "Y" if r.fx_flag else "",
        ])
    _autofit(ws)

    # Sheet 2 - Filtered Kept (mirrors End Result 2024 tab shape)
    ws2 = wb.create_sheet(out_cfg["sheet_kept"])
    ws2.append(["PO No.", "Supplier Name", "Amount", "Description"])
    for d in decisions:
        if d.decision == "keep":
            ws2.append([d.row.po, d.row.supplier, d.row.amount_eur, d.score.kpmg_category or ""])
    _autofit(ws2)

    # Sheet 3 - Review Queue
    # "Recommendation" is the tool's automatic keep/discard call. "Decision" is left
    # blank for the analyst to enter their final say; those verdicts are the ground-truth
    # labels used to train/tune the system later (see README "Future implementation").
    ws3 = wb.create_sheet(out_cfg["sheet_review"])
    ws3.append(["PO Number", "Supplier", "Raw Category", "Amount (EUR)",
                "KPMG Category", "Similarity", "Confidence", "Recommendation",
                "Decision", "Reason"])
    for d in decisions:
        if d.needs_review:
            s = d.score
            ws3.append([d.row.po, d.row.supplier, d.row.category, d.row.amount_eur,
                        s.kpmg_category or "", round(s.similarity, 3),
                        round(s.confidence, 3), d.decision, "", s.reason])
    _autofit(ws3)

    # Sheet 4 - Pivot Summaries (four tables stacked with titles)
    ws4 = wb.create_sheet(out_cfg["sheet_pivots"])
    row_ptr = 1
    for title, df in pivots.items():
        ws4.cell(row=row_ptr, column=1, value=title)
        row_ptr += 1
        ws4.append(list(df.columns)) if False else None  # keep explicit below
        # header
        for j, col in enumerate(df.columns, start=1):
            ws4.cell(row=row_ptr, column=j, value=str(col))
        row_ptr += 1
        for _, rec in df.iterrows():
            for j, col in enumerate(df.columns, start=1):
                val = rec[col]
                ws4.cell(row=row_ptr, column=j, value=val)
            row_ptr += 1
        row_ptr += 1  # blank spacer between tables
    _autofit(ws4)

    wb.save(path)
